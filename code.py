#import all required libraries
import cv2
from picamera2 import Picamera2
import pytesseract
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import RPi.GPIO as GPIO
import time


#define required functions
def beep():
    #initializing GPIO pins 
	GPIO.setmode(GPIO.BOARD)
	GPIO.setwarnings(False)
	GPIO.setup(5, GPIO.OUT)
    #code for buzzer
	for i in range (5):
		GPIO.output(5, GPIO.HIGH)
		time.sleep(0.125)
		GPIO.output(5, GPIO.LOW)
		time.sleep(0.125)

	GPIO.output(5, GPIO.HIGH)
	time.sleep(0.5)
	GPIO.output(5, GPIO.LOW)
	time.sleep(0.5)

	#GPIO.cleanup()

def img_capt(piCam):
    Flag=False        
    #capturing and saving the image
    while True:
        for i in range (70):
            frame=piCam.capture_array()
            cv2.imshow("Number Plate Imager",frame)
            cv2.waitKey(100)
        img=frame
        Flag=True
        img=cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE)
        cv2.imwrite("Image_of_number_plate.jpg",img)
        del piCam
        break
    cv2.destroyAllWindows()
    
    #returning img if captured else 0
    if Flag:
        return img, "Image is Captured", 1
    else:
        return 0, "Error101\nImage was not captured",0
        
def read_text(img):
    text=""
    # cv2.IMREAD_COLOR
    gray=cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    #converting to gray scale
    thresh=cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
    # Apply thresholding to convert to binary image
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3,3))
    cleaned = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel)
    # Apply morphological operations to remove noise and enhance text
    text = pytesseract.image_to_string(cleaned, config='--psm 11')
    # Extract text using Tesseract OCR
    m=len(text)
    text=text[0:m-2]
    #text='KL 06 X 1980'
    if text:
        return text, 1
    else:
        return "Error102\nNo data found",0
    
def check_text(np):
    #the directory is fixed
    os.chdir('/home/boys/Desktop/Project')
    #flag is assigned
    Flag=False
    #workbook is loaded(an external workbook was created earlier)
    wb=load_workbook('vehicle_number.xlsx')
    #worksheet is loaded
    ws=wb.active
    char1=get_column_letter(1)
    char2=get_column_letter(2)
    #checking for the same number and 40 Rs is deducted
    for i in range (2,5):
        if np==ws[char1+str(i)].value:
            ws[char2+str(i)]=(ws[char2+str(i)].value)-40
            bal=ws[char2+str(i)].value
            Flag=True
            wb.save('vehicle_number.xlsx')
            break
    if Flag:
        detail='You may enter and your account baalance is {}'.format(bal),1
    else:
        detail='You cannot Enter',0
    return detail
        
def gate_open():
    # Set GPIO numbering mode
    GPIO.setmode(GPIO.BOARD)
    GPIO.setwarnings(False)

    # Set pin 11 as an output, and set servo1 as pin 11 as PWM
    GPIO.setup(11,GPIO.OUT)
    GPIO.setup(35,GPIO.IN)
    servo1 = GPIO.PWM(11,50) # Note 11 is pin, 50 = 50Hz pulse

    #start PWM running, but with value of 0 (pulse off)
    servo1.start(0)
    print("Gate is opening")
    flag=True
    time.sleep(2)
    servo1.ChangeDutyCycle(7)

    print("Please Pass through the gate")
    while flag:
        #print(GPIO.input(35))
        if GPIO.input(35)==1:
            
            print("Gate closing in 10sec")
            time.sleep(10)
            servo1.ChangeDutyCycle(2)
            time.sleep(0.5)
            servo1.ChangeDutyCycle(0)
            flag=False
            

    #Clean things up at the end
    servo1.stop()
    time.sleep(3)
    #GPIO.cleanup()
    return 1



#Arrives to main function

#initializing GPIO pins 
GPIO.setmode(GPIO.BOARD)
GPIO.setwarnings(False)
GPIO.setup(35, GPIO.IN)

#initializing picam
piCam=Picamera2()
piCam.preview_configuration.main.size=(250,640)
#640,360
#1280,720
piCam.preview_configuration.main.format="RGB888"
piCam.preview_configuration.align()
piCam.configure("preview")
piCam.start()

while True:
    if GPIO.input(35)==0:
        img, statement, img_value=img_capt(piCam)                       #captures image
        print(statement)
        if img_value:
            number_plate, value=read_text(img)                     #reads text in image
            print(number_plate)
            if value:
                result,check=check_text(number_plate)              #checks if the data is correct
                print(result)
                if check:
                    gate_value=gate_open()                         #opens gate
                else:
                    print("Cannot proceed further")
                    beep()
            else:
                print("Cannot proceed further")
                beep()
        else:
            print("Cannot proceed further")
            beep()




------------------------------------------------------------------------------------------------------------------------------------



#importing required libraries
import RPi.GPIO as GPIO
import time

#initializing GPIO pins
GPIO.setmode(GPIO.BOARD)
GPIO.setwarnings(False)

#giving pins input and output values as per use
GPIO.setup(29, GPIO.IN)
GPIO.setup(7, GPIO.OUT)
GPIO.setup(31, GPIO.IN)
GPIO.setup(15, GPIO.OUT)
GPIO.setup(33, GPIO.IN)
GPIO.setup(13, GPIO.OUT)

#checking if car is present
while True:
	val1=GPIO.input(29)
	if val1!=0:
		GPIO.output(7,GPIO.HIGH)
		time.sleep(0.1)
	else:
		GPIO.output(7,GPIO.LOW)
		time.sleep(0.1)
	val2=GPIO.input(31)
	if val2!=0:
		GPIO.output(15,GPIO.HIGH)
		time.sleep(0.1)
	else:
		GPIO.output(15,GPIO.LOW)
		time.sleep(0.1)
	val3=GPIO.input(33)
	if val3!=0:
		GPIO.output(13,GPIO.HIGH)
		time.sleep(0.1)
	else:
		GPIO.output(13,GPIO.LOW)
		time.sleep(0.1)
		
